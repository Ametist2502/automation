from __future__ import annotations

import csv
import re
from pathlib import Path

import pytest
from openpyxl import load_workbook

from main import build_comment, load_flat_file, main, normalize_header, normalize_headers


def test_normalize_header() -> None:
    assert normalize_header("  Sales Amount ") == "salesamount"


def test_normalize_headers_duplicate_raises() -> None:
    with pytest.raises(ValueError):
        normalize_headers(["Sale Amount", "saleamount"])


def test_build_comment() -> None:
    result = build_comment(
        value="120",
        column="sales",
        row_number=2,
        template="{column}={value} at {row_number}",
    )
    assert result == "sales=120 at 2"


def test_load_flat_file_csv_normalize_header(tmp_path: Path) -> None:
    csv_file = tmp_path / "input.csv"
    csv_file.write_text("  Region Name , Sales Amount\nHN,120\n", encoding="utf-8-sig")

    columns, rows, source_info = load_flat_file(csv_file)

    assert columns == ["regionname", "salesamount"]
    assert rows == [{"regionname": "HN", "salesamount": "120"}]
    assert source_info in {"utf-8-sig", "utf-8"}


def test_main_csv_creates_output_file_and_log(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
    capsys: pytest.CaptureFixture[str],
) -> None:
    input_file = tmp_path / "input.csv"
    input_file.write_text(
        " Region , Sales Amount\nHN,120\nHCM,200\n",
        encoding="utf-8",
    )

    output_file = tmp_path / "output.csv"
    log_file = tmp_path / "run.log"
    monkeypatch.setenv("AUTO_COMMENT_LOG_FILE", str(log_file))

    answers = iter(
        [
            str(input_file),
            "Region={value}",
            "Sales={value}",
            "comment",
            " || ",
            str(output_file),
        ]
    )

    monkeypatch.setattr("builtins.input", lambda _prompt: next(answers))

    main()

    out = capsys.readouterr().out
    assert "Completed!" in out
    assert output_file.exists()
    assert log_file.exists()

    with output_file.open("r", newline="", encoding="utf-8-sig") as f:
        rows = list(csv.DictReader(f))

    assert rows[0]["comment"] == "Region=HN || Sales=120"
    assert rows[1]["comment"] == "Region=HCM || Sales=200"

    log_text = log_file.read_text(encoding="utf-8")
    assert "-INFO: Application started" in log_text
    first_line = log_text.splitlines()[0]
    assert re.match(r"\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}-[A-Z]+: .+", first_line)


def test_main_xlsx_creates_output_file(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
    capsys: pytest.CaptureFixture[str],
) -> None:
    input_file = tmp_path / "input.xlsx"
    output_file = tmp_path / "output.xlsx"

    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.append([" Product Name ", "Qty Sold"])
    ws.append(["Pen", 10])
    ws.append(["Book", 5])
    wb.save(input_file)

    answers = iter(
        [
            str(input_file),
            "Product={value}",
            "Qty={value}",
            "comment",
            " | ",
            str(output_file),
        ]
    )

    monkeypatch.setattr("builtins.input", lambda _prompt: next(answers))

    main()

    out = capsys.readouterr().out
    assert "Completed!" in out
    assert output_file.exists()

    wb_out = load_workbook(output_file)
    ws_out = wb_out.active

    headers = [cell.value for cell in next(ws_out.iter_rows(min_row=1, max_row=1))]
    assert headers == ["productname", "qtysold", "comment"]

    row2 = [cell.value for cell in next(ws_out.iter_rows(min_row=2, max_row=2))]
    row3 = [cell.value for cell in next(ws_out.iter_rows(min_row=3, max_row=3))]

    assert row2[2] == "Product=Pen | Qty=10"
    assert row3[2] == "Product=Book | Qty=5"
