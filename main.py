from __future__ import annotations

import csv
import logging
import os
from pathlib import Path

from openpyxl import Workbook, load_workbook

LOGGER_NAME = "auto_comment"
DEFAULT_LOG_FILE = "auto_comment.log"


def setup_logging(log_file: Path) -> logging.Logger:
    logger = logging.getLogger(LOGGER_NAME)
    logger.setLevel(logging.INFO)
    logger.propagate = False

    logger.handlers.clear()

    log_file.parent.mkdir(parents=True, exist_ok=True)
    file_handler = logging.FileHandler(log_file, encoding="utf-8")
    file_handler.setFormatter(logging.Formatter("%(asctime)s-%(levelname)s: %(message)s", "%Y-%m-%d %H:%M:%S"))

    logger.addHandler(file_handler)
    return logger


def normalize_header(header: str) -> str:
    return "".join(header.strip().lower().split())


def normalize_headers(headers: list[str]) -> list[str]:
    normalized = [normalize_header(h) for h in headers]

    if any(not h for h in normalized):
        raise ValueError("Header is empty after normalization. Please check the input file.")

    duplicates = {h for h in normalized if normalized.count(h) > 1}
    if duplicates:
        duplicated = ", ".join(sorted(duplicates))
        raise ValueError(f"Duplicate headers after normalization: {duplicated}")

    return normalized


def to_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def read_csv_with_fallback(file_path: Path) -> tuple[list[str], list[dict[str, str]], str]:
    encodings = ["utf-8-sig", "utf-8", "cp1258", "latin-1"]

    last_error: Exception | None = None
    for encoding in encodings:
        try:
            with file_path.open("r", newline="", encoding=encoding) as f:
                reader = csv.DictReader(f)
                if not reader.fieldnames:
                    raise ValueError("Header row was not found in CSV file.")

                rows = list(reader)
                return reader.fieldnames, rows, encoding
        except Exception as err:  # noqa: BLE001
            last_error = err

    raise RuntimeError(f"Could not read CSV file. Last error: {last_error}")


def read_xlsx(file_path: Path) -> tuple[list[str], list[dict[str, str]], str]:
    workbook = load_workbook(filename=file_path, data_only=True)
    sheet = workbook.active

    data = list(sheet.iter_rows(values_only=True))
    if not data:
        raise ValueError("XLSX file is empty.")

    raw_headers = [to_text(v) for v in data[0]]
    if not any(raw_headers):
        raise ValueError("Header row was not found in XLSX file.")

    headers = raw_headers
    rows: list[dict[str, str]] = []

    for row_values in data[1:]:
        row_dict: dict[str, str] = {}
        for idx, header in enumerate(headers):
            if not header:
                continue
            value = row_values[idx] if idx < len(row_values) else ""
            row_dict[header] = to_text(value)
        rows.append(row_dict)

    return headers, rows, "xlsx"


def normalize_table(headers: list[str], rows: list[dict[str, str]]) -> tuple[list[str], list[dict[str, str]]]:
    normalized_headers = normalize_headers(headers)
    mapping = {original: normalized for original, normalized in zip(headers, normalized_headers, strict=True)}

    normalized_rows: list[dict[str, str]] = []
    for row in rows:
        normalized_row: dict[str, str] = {}
        for original_header, normalized_header in mapping.items():
            value = row.get(original_header, "")
            normalized_row[normalized_header] = to_text(value)
        normalized_rows.append(normalized_row)

    return normalized_headers, normalized_rows


def load_flat_file(file_path: Path) -> tuple[list[str], list[dict[str, str]], str]:
    suffix = file_path.suffix.lower()

    if suffix == ".csv":
        headers, rows, source_info = read_csv_with_fallback(file_path)
    elif suffix == ".xlsx":
        headers, rows, source_info = read_xlsx(file_path)
    else:
        raise ValueError("Only .csv and .xlsx files are supported.")

    normalized_headers, normalized_rows = normalize_table(headers, rows)
    return normalized_headers, normalized_rows, source_info


def write_csv(file_path: Path, headers: list[str], rows: list[dict[str, str]]) -> None:
    with file_path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=headers)
        writer.writeheader()
        writer.writerows(rows)


def write_xlsx(file_path: Path, headers: list[str], rows: list[dict[str, str]]) -> None:
    workbook = Workbook()
    sheet = workbook.active

    sheet.append(headers)
    for row in rows:
        sheet.append([row.get(h, "") for h in headers])

    workbook.save(file_path)


def save_flat_file(file_path: Path, headers: list[str], rows: list[dict[str, str]]) -> None:
    suffix = file_path.suffix.lower()

    if suffix == ".csv":
        write_csv(file_path, headers, rows)
    elif suffix == ".xlsx":
        write_xlsx(file_path, headers, rows)
    else:
        raise ValueError("Output file must use .csv or .xlsx extension.")


def ask_input(prompt: str, default: str | None = None, allow_empty: bool = False, strip_value: bool = True) -> str:
    while True:
        if default is not None:
            raw = input(f"{prompt} [{default}]: ")
            if strip_value:
                raw = raw.strip()
            return raw or default

        raw = input(f"{prompt}: ")
        if strip_value:
            raw = raw.strip()
        if raw or allow_empty:
            return raw

        print("Value cannot be empty. Please try again.")


def print_columns(columns: list[str]) -> None:
    print("\nAvailable columns after header normalization:")
    for idx, col in enumerate(columns, start=1):
        print(f"  {idx}. {col}")


def build_comment(value: str, column: str, row_number: int, template: str) -> str:
    return template.format(
        value=value,
        column=column,
        row_number=row_number,
    )


def ask_column_templates(columns: list[str]) -> dict[str, str]:
    print("\nEnter comment template for each column.")
    print("Leave blank to skip a column.")
    print("Supported placeholders: {column}, {value}, {row_number}\n")

    templates: dict[str, str] = {}
    for col in columns:
        template = ask_input(f"Template for column '{col}'", allow_empty=True)
        if template:
            templates[col] = template

    return templates


def main() -> None:
    log_file = Path(os.getenv("AUTO_COMMENT_LOG_FILE", DEFAULT_LOG_FILE))
    logger = setup_logging(log_file)

    print("=== AUTO COMMENT FLAT FILE (CSV/XLSX) ===")
    logger.info("Application started")

    input_path = Path(ask_input("Enter input file path (.csv/.xlsx)", "input.csv"))
    logger.info("Input file provided: %s", input_path)

    if not input_path.exists() or not input_path.is_file():
        print(f"File not found: {input_path}")
        logger.error("File not found: %s", input_path)
        return

    try:
        columns, rows, source_info = load_flat_file(input_path)
    except Exception as err:  # noqa: BLE001
        print(f"Failed to read file: {err}")
        logger.exception("Failed to read file: %s", err)
        return

    if not rows:
        print("The file contains no data rows.")
        logger.warning("No data rows in file: %s", input_path)
        return

    print_columns(columns)
    logger.info("Loaded %d rows with %d columns", len(rows), len(columns))

    column_templates = ask_column_templates(columns)
    if not column_templates:
        print("No template was provided. Nothing to generate.")
        logger.warning("No templates provided by user")
        return

    comment_column = ask_input("New comment column name", "comment")
    separator = ask_input("Separator between comments", " | ", strip_value=False)
    logger.info("Comment column: %s, templates configured: %d", comment_column, len(column_templates))

    default_output = f"{input_path.stem}_commented{input_path.suffix.lower()}"
    output_path = Path(ask_input("Output file path", default_output))

    if comment_column not in columns:
        columns.append(comment_column)

    for row_number, row in enumerate(rows, start=1):
        comments: list[str] = []

        for col, template in column_templates.items():
            value = row.get(col, "")
            comments.append(
                build_comment(
                    value=value,
                    column=col,
                    row_number=row_number,
                    template=template,
                )
            )

        row[comment_column] = separator.join(comments)

    try:
        save_flat_file(output_path, columns, rows)
    except Exception as err:  # noqa: BLE001
        print(f"Failed to write file: {err}")
        logger.exception("Failed to write output file: %s", err)
        return

    print("\nCompleted!")
    print(f"- Loaded: {input_path} ({source_info})")
    print(f"- Saved: {output_path}")
    print(f"- Commented rows: {len(rows)}")
    print(f"- Log file: {log_file}")

    logger.info("Completed successfully. Output file: %s. Rows processed: %d", output_path, len(rows))


if __name__ == "__main__":
    main()
